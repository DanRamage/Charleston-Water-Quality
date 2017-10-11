import sys
sys.path.append('../commonfiles/python')
import os
import logging.config
import optparse
import ConfigParser
from datetime import datetime, timedelta
import time
from wqDatabase import wqDB
import csv

from pyoos.collectors.usgs.usgs_rest import UsgsRest
from pyoos.collectors.coops.coops_sos import CoopsSos


from pytz import timezone
import json
from wq_sites import wq_sample_sites

from unitsConversion import uomconversionFunctions
from build_tide_file import create_tide_data_file_mp
from wqXMRGProcessing import wqXMRGProcessing
from wq_output_results import wq_sample_data,wq_samples_collection


usgs_sites = ['021720709', '021720869', '021720710', '021720698','0217206935','021720677']
usgs_code_to_obs = {
  '00010': { 'usgs_obs_name': 'water_temperature',
             'xenia_obs_name': 'water_temperature',
             'usgs_units': 'celsius',
             'xenia_units': 'celsius'
          },
  '00065': { 'usgs_obs_name':'gage_height',
             'xenia_obs_name': 'gage_height',
             'usgs_units': 'ft',
             'xenia_units': 'm'
             },
  '00095': { 'usgs_obs_name':'specific_conductance',
             'xenia_obs_name': 'water_conductivity',
             'usgs_units': 'uS_cm-1',
             'xenia_units': 'uS_cm-1'
             },
  '00300': {'usgs_obs_name':'dissolved_oxygen',
            'xenia_obs_name': 'oxygen_concentration',
            'usgs_units': 'mg_L-1',
            'xenia_units': 'mg_L-1'
            },

  '00480': { 'usgs_obs_name':'salinity',
             'xenia_obs_name': 'salinity',
             'usgs_units': 'psu',
             'xenia_units': 'psu'
             }
}

nos_sites = ['8665530']
nos_obs = {
  "sea_water_temperature": {
                             "units": "celsius",
                             "xenia_name": "water_temperature",
                             "xenia_units": "celsius"
                           },
  "wind_speed": {
                  "units": "m_s-1",
                  "xenia_name": "wind_speed",
                  "xenia_units": "m_s-1"

                },
  "wind_from_direction": {
    "units": "degrees_true",
    "xenia_name": "wind_from_direction",
    "xenia_units": "degrees_true"

  }
}

tide_stations = ['8665530']

def flatten_element(p):
  rd = {'time': p.time}
  for m in p.members:
    rd[m['standard']] = { 'value': m['value'], 'units': m['unit'] }

  return rd


def get_usgs_data(site, dates, units_converter):
  start_time = time.time()
  logger = logging.getLogger('chs_historical_logger')
  logger.debug("Starting get_usgs_data")

  station_data = {}
  usgs_rest = UsgsRest()
  utc_tz = timezone('UTC')
  for rec_date in dates:
    usgs_rest.clear()
    end_date = rec_date.astimezone(utc_tz) + timedelta(hours=24)
    start_date = rec_date.astimezone(utc_tz) - timedelta(hours=24)
    logger.debug("Query site: %s for date: %s Begin: %s End: %s" % (site, rec_date, start_date, end_date))

    #usgs_rest.filter(features=[site], start=start_date, end=rec_date, variables=['00010','00010','00065','00095','00095','00300','00300','00480','00480'])
    retries = 3
    for i in range(0, retries):
      try:
        usgs_rest.filter(features=[site], start=start_date, end=end_date)
        results = usgs_rest.collect()
      except Exception as e:
        logger.error("Site: %s Date: %s query num: %d problem." % (site, rec_date, i))
        logger.exception(e)
      else:
        break
    if len(results.elements):
      sta = results.elements[0]
      #flat_results = map(flatten_element, sta.elements)
      """
      for rec in sta.elements:
        if rec.time not in station_data:
          station_data[rec.time] = {}
        time_rec = station_data[rec.time]
        for member in rec.members:
          obs_name = member['standard']
          if obs_name in usgs_code_to_obs:
            obs_name = usgs_code_to_obs[obs_name]
          if obs_name not in time_rec:
            time_rec[obs_name] = []
          obs_rec = time_rec[obs_name]
          obs_rec.append({'value': member['value'],
                          'units': member['unit']})
      """
      for rec in sta.elements:
        for ndx, member in enumerate(rec.members):
          obs_name = member['standard']
          if obs_name in usgs_code_to_obs:
            obs_name = usgs_code_to_obs[obs_name]['xenia_obs_name']
          if obs_name not in station_data:
            station_data[obs_name] = []

          station_rec = station_data[obs_name]
          obs_rec = usgs_code_to_obs[member['standard']]
          if obs_rec['usgs_units'] == 'ft':
            i = 0
          try:
            obs_val = float(member['value'])
            obs_val = units_converter.measurementConvert(obs_val,
                                                         obs_rec['usgs_units'],
                                                         obs_rec['xenia_units'])
          except ValueError as e:
            logger.exception(e)
          else:
            if obs_val is None:
              obs_val = member['value']

            station_rec.append({'date_time': rec.time.strftime('%Y-%m-%dT%H:%M:%S'),
                                'value': obs_val,
                                'units': obs_rec['xenia_units'],
                                'latitude': rec.location.y,
                                'longitude': rec.location.x,
                                's_order': ndx + 1})
      #Sort based on the date field
      for param in station_data:
        station_data[param].sort(key=lambda x: x['date_time'])

  logger.debug("Finished get_usgs_data. Queried %d dates in %f seconds" % (len(dates), time.time() - start_time))
  return station_data

def parse_sheet_data(xl_file_name, wq_sites, wq_data_collection):
  from openpyxl import load_workbook

  logger = logging.getLogger('chs_historical_logger')
  logger.debug("Starting parse_sheet_data, parsing file: %s" % (xl_file_name))


  #sample_data = {}
  wb = load_workbook(filename = xl_file_name)

  est_tz = timezone('US/Eastern')


  for ws in wb:
    logger.debug("Worksheet: %s" % (ws))
    row_num = 0
    date_list = []
    for row in ws.iter_rows():
      #logger.debug("Row: %d" % (row_num))
      #Cell A is the site name, then subsequent cells are the dates on first row and sample data on other rows.
      col_num = 0
      current_site = None
      for col in row:
        #logger.debug("Column: %d" % (col_num))
        if row_num == 0:
          #Skip site column
          if col_num > 0:
            if col.value is not None:
              #Set correct timezone
              est_time = est_tz.localize(col.value)
              date_list.append(est_time)
        else:
          if col_num == 0:
            current_site = None
            if col.value != None:
              cleaned_sample_site_name = col.value.strip(' ')
              for site in wq_sites:
                if cleaned_sample_site_name.find(site.name) != -1:
                  current_site = site.name
              """
              for site in wq_sites:
                if cleaned_sample_site_name.find(site.name) != -1:
                  current_site = site.name
                  if current_site not in sample_data:
                    sample_data[current_site] = []
                  break
              """
            else:
              break
          else:
            if current_site == None:
              break
            try:
              value = None
              if col.value is not None:
                value = float(col.value)
              else:
                break
            except ValueError as e:
              try:
                value = float(col.value.replace('<', ''))
              except ValueError as e:
                value = None
            try:
              logger.debug("Site: %s Date: %s Value: %s" % (current_site, date_list[col_num - 1], value))
              wq_sample_rec = wq_sample_data()
              wq_sample_rec.station = current_site
              wq_sample_rec.date_time = date_list[col_num-1]
              wq_sample_rec.value = value
              wq_data_collection.append(wq_sample_rec)

              #sample_data[current_site].append({
              #  'site_name': current_site,
              #  'sample_date': date_list[col_num-1],
              #  'value': value
              #})
            except Exception as e:
              logger.exception(e)
        col_num += 1
      row_num +=1
  #for site in sample_data:
  #  if site is not None:
  #    sample_data[site].sort(key=lambda x: x['sample_date'], reverse=True)

  if logger:
    logger.debug("Finished parse_sheet_data")


def parse_dhec_sheet_data(xl_file_name, wq_data_collection):
  from xlrd import xldate
  import xlrd


  station_to_site_map = {
    'AR2': 'Brittlebank Park',
    'JIC2': 'James Island Creek 2',
    'JIC1': 'James Island Creek 1',
    'SC1': 'Shem Creek 1',
    'SC2': 'Shem Creek 2',
    'SC3': 'Shem Creek 3',
    'FB1': 'Folly Beach'}
  logger = logging.getLogger('chs_historical_logger')
  logger.debug("Starting parse_dhec_sheet_data, parsing file: %s" % (xl_file_name))

  wb = xlrd.open_workbook(filename = xl_file_name)

  est_tz = timezone('US/Eastern')
  #utc_tz = timezone('UTC')
  sample_date = None
  try:
    sheet = wb.sheet_by_name('Results')
  except Exception as e:
    logger.exception(e)
  else:
    row_headers = []
    results_ndx = \
    station_ndx = \
    date_ndx = \
    parameter_ndx = \
    time_ndx = None
    for row_ndx,data_row in enumerate(sheet.get_rows()):
      if row_ndx != 0:
        try:
          wq_sample_rec = wq_sample_data()
          if data_row[parameter_ndx].value == "Enterococci":
            station_name = data_row[station_ndx].value.strip()
            if station_name in station_to_site_map:
              wq_sample_rec.station = station_to_site_map[station_name]
              try:
                date_val = xlrd.xldate.xldate_as_datetime(data_row[date_ndx].value, wb.datemode)
              except Exception as e:
                try:
                  date_val = datetime.strptime(data_row[date_ndx].value, "%Y-%m-%d")
                except Exception as e:
                  logger.error("Date format error on line: %d" % (row_ndx))
                  logger.exception(e)
                  break
              try:
                time_val = xldate.xldate_as_datetime(data_row[time_ndx].value, False)
                #time_val = datetime.strptime(data_row[time_ndx].value, "%H%M")
              except Exception as e:
                val = data_row[time_ndx].value
                try:
                  time_val = datetime.strptime(str(val), "%H%M")
                except Exception as e:
                  logger.error("Time format error on line: %d" % (row_ndx))
                  time_val = datetime.strptime('00:00:00', '%H:%M:%S')

              wq_sample_rec.date_time = (est_tz.localize(datetime.combine(date_val.date(), time_val.time())))
              #wq_sample_rec.date_time = (est_tz.localize(datetime.combine(date_val.date(), time_val.time()))).astimezone(utc_tz)
              wq_sample_rec.value = data_row[results_ndx].value
              logger.debug("Site: %s Date: %s Value: %s" % (wq_sample_rec.station,
                                                            wq_sample_rec.date_time,
                                                            wq_sample_rec.value))
              if sample_date is None or date_val > sample_date:
                sample_date = date_val
              wq_data_collection.append(wq_sample_rec)
        except Exception as e:
          logger.error("Error found on row: %d" % (row_ndx))
          logger.exception(e)
      else:
        #Copy the header names out
        for cell in data_row:
          row_headers.append(cell.value)
        station_ndx = row_headers.index('Station')
        date_ndx = row_headers.index('Date')
        time_ndx = row_headers.index('Time')
        results_ndx = row_headers.index('Result')
        parameter_ndx = row_headers.index("Parameter")

  return sample_date

def process_usgs_data(**kwargs):
  logger = logging.getLogger('chs_historical_logger')

  out_directory = kwargs['output_directory']
  all_dates = kwargs['all_dates']
  db_obj = kwargs['db_obj']
  units_converter = kwargs['units_converter']
  #Go grab the data from the web service.
  if kwargs.get('query_usgs', False):
    for usgs_site in usgs_sites:
      # station_data = get_usgs_data(usgs_site, [datetime.strptime('2016-10-26T00:00:00', "%Y-%m-%dT%H:%M:%S")])
      station_data = get_usgs_data(usgs_site, all_dates, units_converter)
      try:
        out_file = os.path.join(out_directory, "%s.json" % (usgs_site))
        with open(out_file, "w") as out_file_obj:
          out_file_obj.write(json.dumps(station_data, sort_keys=True, indent=2, separators=(',', ': ')))
      except(IOError, Exception) as e:
        logger.exception(e)
  #Process pre-fetched data.
  else:
    row_entry_date = datetime.now()
    for usgs_site in usgs_sites:
      try:
        out_file = os.path.join(out_directory, "%s.json" % (usgs_site))
        logger.debug("Processing file: %s" % (out_file))
        with open(out_file, "r") as out_file_obj:
          platform_handle = "usgs.%s.wq" % (usgs_site)
          json_data = json.load(out_file_obj)
          #Get the observations
          obs_keys = json_data.keys()
          #Check if platform exists, if not then add platform and observations.
          if db_obj.platformExists(platform_handle) == -1:
            obs_list = []
            for obs_name in obs_keys:
              #Get the units and s order from data.
              data_list = json_data[obs_name]
              if len(data_list):
                rec = data_list[0]
                obs_list.append({'obs_name':obs_name,
                                 'uom_name': rec['units'],
                                 's_order': rec['s_order']})
            db_obj.buildMinimalPlatform(platform_handle, obs_list)
          #Add records into db
          for obs_name in obs_keys:
            data_list = json_data[obs_name]
            for data_rec in data_list:
              if len(data_list):
                try:
                  logger.debug("Station: %s Obs: %s Date: %s Val: %s s_order: %d"%\
                               (platform_handle,
                                obs_name,
                                data_rec['date_time'],
                                data_rec['value'],
                                data_rec['s_order']))

                  db_obj.addMeasurement(obs_name,
                                        data_rec['units'],
                                        platform_handle,
                                        data_rec['date_time'],
                                        data_rec['latitude'],
                                        data_rec['longitude'],
                                         0,
                                         [float(data_rec['value'])],
                                         sOrder=data_rec['s_order'],
                                         autoCommit=True,
                                         rowEntryDate=row_entry_date)
                except Exception as e:
                  logger.exception(e)
      except(IOError, Exception) as e:
        logger.exception(e)

  return

def get_nos_data(site, dates, units_coverter, db_obj):
  start_time = time.time()
  logger = logging.getLogger('chs_historical_logger')
  logger.debug("Starting get_nos_data")

  row_entry_date = datetime.now()
  utc_tz = timezone('UTC')
  eastern_tz= timezone('US/Eastern')

  platform_handle = 'nos.%s.met' % (site)
  if db_obj.platformExists(platform_handle) == -1:
    obs_list = []
    for single_obs in nos_obs:
      obs_list.append({'obs_name': nos_obs[single_obs]['xenia_name'],
                       'uom_name': nos_obs[single_obs]['xenia_units'],
                       's_order': 1})
    db_obj.buildMinimalPlatform(platform_handle, obs_list)

  nos_query = CoopsSos()
  #dates.sort(reverse=True)
  for rec_date in dates:
    logger.debug("Query site: %s for date: %s" % (site, rec_date))
    nos_query.clear()
    utc_end_date = rec_date.astimezone(utc_tz) + timedelta(hours=24)
    start_date = rec_date.astimezone(utc_tz) - timedelta(hours=24)

    for single_obs in nos_obs:
      obs_type = nos_obs[single_obs]['xenia_name']
      uom_type = nos_obs[single_obs]['xenia_units']
      s_order = 1

      nos_query.filter(features=[site], start=start_date, end=utc_end_date, variables=[single_obs])
      try:
        #results = nos_query.collect()
        response = nos_query.raw(responseFormat="text/csv")
      except Exception as e:
        logger.exception(e)
      else:
        csv_reader = csv.reader(response.split('\n'), delimiter=',')
        line_cnt = 0
        for row in csv_reader:
          if line_cnt > 0 and len(row):
            obs_date = datetime.strptime(row[4], '%Y-%m-%dT%H:%M:%SZ')
            obs_val = float(row[5])
            logger.debug("Adding obs: %s(%s) Date: %s Value: %s S_Order: %d" %\
                         (obs_type, uom_type, obs_date, obs_val, s_order))

            if not db_obj.addMeasurement(obs_type,
                                    uom_type,
                                    platform_handle,
                                    obs_date.strftime('%Y-%m-%dT%H:%M:%S'),
                                    float(row[2]),
                                    float(row[3]),
                                    0,
                                    [obs_val],
                                    sOrder=s_order,
                                    autoCommit=True,
                                    rowEntryDate=row_entry_date ):
              logger.error(db_obj.lastErrorMsg)


          line_cnt += 1

  logger.debug("Finished get_nos_data in %f seconds" % (time.time() - start_time))

  return
def process_nos_data(**kwargs):
  out_directory = kwargs['output_directory']
  all_dates = kwargs['all_dates']
  db_obj = kwargs['db_obj']
  units_converter = kwargs['units_converter']

  if kwargs.get('query_ndbc', False):
    for site in nos_sites:
      get_nos_data(site, all_dates, units_converter, db_obj)
  else:
    i=0
  return

def process_tide_data(**kwargs):
  start_time = time.time()
  logger = logging.getLogger('chs_historical_logger')
  logger.debug("Starting process_tide_data")

  out_directory = kwargs['output_directory']
  all_dates = kwargs['all_dates']
  db_obj = kwargs['db_obj']
  units_converter = kwargs['units_converter']
  log_conf_file = kwargs['log_config_file']
  eastern_tz = timezone('US/Eastern')
  tide_dates = []
  for tide_station in tide_stations:
    for date_rec in all_dates:
      #Add 24 hours since we want to make sure we have +/- 24 hours around our date. This
      #way we can have enough data to use if we want the sample times starting at midnight
      #or we want to use the actual sample time. Instead of getting the data for each
      #time for a given sample day, just do a more coarse approach.
      #tide_date = date_rec + timedelta(hours=24)
      #tide_date = tide_date.replace(hour=0, minute=0, second=0)
      tide_dates.append(date_rec)

  tide_output_file = os.path.join(out_directory, "%s.csv" % (tide_station))
  create_tide_data_file_mp(tide_station,
                           tide_dates,
                           tide_output_file,
                           4,
                           log_conf_file,
                           True)

  logger.debug("Finished process_tide_data in %f seconds" % (time.time()-start_time))

def process_nexrad_data(**kwargs):
  start_time = time.time()
  logger = logging.getLogger('chs_historical_logger')
  logger.debug("Starting process_tide_data")

  config_file = kwargs['config_file']
  import_directory = kwargs['import_directory']

  xmrg_proc = wqXMRGProcessing(logger=True)
  xmrg_proc.load_config_settings(config_file=config_file)

  logger.info("Importing directory(s): %s" % (import_directory))

  import_dirs = import_directory.split(",")

  for import_dir in import_dirs:
    file_list = os.listdir(import_dir)
    file_list.sort()
    full_path_file_list = [os.path.join(import_dir, file_name) for file_name in file_list]
    xmrg_proc.import_files(full_path_file_list)

  logger.debug("Finished process_nexrad_data in %f seconds" % (time.time()-start_time))
  return


def main():
  parser = optparse.OptionParser()
  parser.add_option("--ConfigFile", dest="config_file", default=None,
                    help="INI Configuration file." )
  parser.add_option("--WaterKeeperHistoryFile", dest="water_keeper_excel_file", default=None,
                    help="Excel File with historical etcoc data." )
  parser.add_option("--DHECHistoryFile", dest="dhec_excel_file", default=None,
                    help="DHEC Excel File with historical etcoc data." )

  parser.add_option("--OutputDirectory", dest="out_dir", default='./')
  parser.add_option("--CreateCurrentSampleFile", dest="create_current", default=False,
                    help="If set, this will create the current the stations json file with the latest data.")
  parser.add_option("--CreateStationsFile", dest="create_stations", default=False,
                    help="If set, this will create or update the individual station json file with the data from latest email.")
  parser.add_option("--DateFile", dest="date_file", default=None,
                    help="If provided, full file path for a list of dates processed.")

  parser.add_option("--ProcessUSGSData", dest="process_usgs_data", action="store_true",
                    help="If set this will process USGS data, either by querying the server or processing the json files.")
  parser.add_option("--QueryUSGS", dest="query_usgs", action="store_true",
                    help="If set this will query the USGS web service to get the historical data.")
  parser.add_option("--ProcessNDBCData", dest="process_ndbc_data", action="store_true",
                    help="If set this will process NDBC data, either by querying the server or processing the json files.")
  parser.add_option("--QueryNDBC", dest="query_ndbc", action="store_true",
                    help="If set this will query the NDBC web service to get the historical data.")
  parser.add_option("--ProcessTideData", dest="process_tide_data", action="store_true",
                    help="If set this will process tide data, either by querying the server or processing the json files.")
  parser.add_option("--QueryTide", dest="query_tide", action="store_true",
                    help="If set this will query the tide web service to get the historical data.")
  parser.add_option("--ProcessNexrad", dest="process_nexrad", action="store_true",
                    help="If set this will import NEXRAD data from directory list given in --NEXRADDirectory")
  parser.add_option("--NEXRADDirectoryList", dest="nexrad_directory_list",
                    help="If we are querying NEXRAD data, this is a list of directories to import data into.")

  (options, args) = parser.parse_args()

  if options.config_file is None:
    parser.print_help()
    sys.exit(-1)

  try:
    config_file = ConfigParser.RawConfigParser()
    config_file.read(options.config_file)
  except Exception, e:
    raise
  else:
    logger = None
    try:
      logConfFile = config_file.get('logging', 'config_file')
      boundaries_location_file = config_file.get('boundaries_settings', 'boundaries_file')
      sites_location_file = config_file.get('boundaries_settings', 'sample_sites')

      historical_db_name = config_file.get("historic_database", "name")
      units_file = config_file.get("units_conversion", "config_file")
      if logConfFile:
        logging.config.fileConfig(logConfFile)
        logger = logging.getLogger('chs_historical_logger')
        logger.info("Log file opened.")
    except ConfigParser.Error, e:
      if logger:
        logger.exception(e)
    else:
      units_conversion = uomconversionFunctions(units_file)
      historic_db = wqDB(historical_db_name, 'chs_historical_logger')

      wq_sites = wq_sample_sites()
      wq_sites.load_sites(file_name=sites_location_file, boundary_file=boundaries_location_file)

      chs_results = wq_samples_collection()

      if options.water_keeper_excel_file is not None:
        chs_results = parse_sheet_data(options.excel_file, wq_sites, chs_results)

      if options.dhec_excel_file is not None:
        file_list = options.dhec_excel_file.split(',')
        for file in file_list:
          parse_dhec_sheet_data(file, chs_results)

      #Build list of all dates.
      all_dates = []
      est_tz = timezone('US/Eastern')
      for site in chs_results:
        logger.debug("Site: %s getting dates" % (site))
        site_data = chs_results[site]
        for rec in site_data:
          #if rec['sample_date'] not in all_dates:
          if rec.date_time.date() not in all_dates:
            logger.debug("Site: %s adding date: %s" % (site, rec.date_time.date()))
            all_dates.append(rec.date_time.date())
            #all_dates.append(rec['sample_date'])
      for ndx,date in enumerate(all_dates):
        date = est_tz.localize(datetime.combine(date, datetime.min.time()))
        all_dates[ndx] = date

      all_dates.sort()
      #Getting/Processing USGS data?
      if options.process_usgs_data:
        process_usgs_data(output_directory=options.out_dir,
                          query_usgs=options.query_usgs,
                          all_dates=all_dates,
                          db_obj=historic_db,
                          units_converter=units_conversion)
      if options.process_ndbc_data:
        process_nos_data(output_directory=options.out_dir,
                          query_ndbc=options.query_ndbc,
                          all_dates=all_dates,
                          db_obj=historic_db,
                          units_converter=units_conversion)
      if options.process_tide_data:
        #For tides, we want to use the samples date and time.
        tide_dates = []
        last_date = None
        for site in chs_results:
          logger.debug("Site: %s getting dates" % (site))
          site_data = chs_results[site]
          for rec in site_data:
            if (last_date is None or last_date != rec.date_time):
              last_date = est_tz.localize(datetime.combine(rec.date_time.date(), datetime.min.time()))
              if last_date not in tide_dates:
                tide_dates.append(last_date)

            logger.debug("Site: %s adding date: %s" % (site, rec.date_time))
            tide_dates.append(rec.date_time)

        process_tide_data(output_directory=options.out_dir,
                          query_tide=options.query_tide,
                          all_dates=tide_dates,
                          db_obj=historic_db,
                          units_converter=units_conversion,
                          log_config_file=logConfFile)

        if options.process_nexrad:
          process_nexrad_data(output_directory=options.out_dir,
                              config_file=options.config_file,
                              import_directory=options.nexrad_directory_list)
    logger.debug("Closing log.")
if __name__ == "__main__":
  main()
